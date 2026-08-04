# -*- coding: utf-8 -*-
"""MES 일일 기록 샘플 구조 분석기 (7/20 샘플 수령 시 1순위 실행).

사용법:
    python scripts/mes_spike.py "..\\MES다운로드\\2026-07-20"   # 프로젝트 루트 기준 — 드라이브 문자 무관
    python scripts/mes_spike.py <파일.xlsx>            # 단일 파일도 가능

출력: 파일별 시트 목록, 시트별 크기, 상단 데이터 미리보기, 병합영역·이미지 개수.
목적: MES 내보내기가 (a)표 그리드형인지 (b)기준서형 셀맵인지 판정 → 파서 설계.
관련 계획: docs/sq-levelup-2026-10/40_CP주기의무_매핑표.md + 메모리 plan-sq-levelup-audit.
"""
import sys
import os
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')


def inspect(path: str) -> None:
    print('=' * 90)
    print('FILE:', os.path.basename(path), '(%.1f MB)' % (os.path.getsize(path) / 1048576))
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
    except Exception as e:
        print('  열기 실패:', type(e).__name__, str(e)[:200])
        return
    print('  시트 %d개:' % len(wb.sheetnames), wb.sheetnames[:25])
    for name in wb.sheetnames[:6]:
        ws = wb[name]
        imgs = len(getattr(ws, '_images', []))
        merged = len(ws.merged_cells.ranges)
        print('-' * 90)
        print('  [%s] %d행 x %d열 · 병합 %d · 이미지 %d' % (name, ws.max_row, ws.max_column, merged, imgs))
        shown = 0
        for row in ws.iter_rows(min_row=1, max_row=25, max_col=16, values_only=True):
            vals = ['' if v is None else str(v)[:14] for v in row]
            if any(vals):
                print('   ', ' | '.join(vals))
                shown += 1
            if shown >= 14:
                break
    wb.close()


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        return
    target = sys.argv[1]
    if os.path.isdir(target):
        files = [os.path.join(target, f) for f in sorted(os.listdir(target)) if f.lower().endswith(('.xlsx', '.xlsm'))]
        if not files:
            print('폴더에 xlsx 파일이 없습니다:', target)
            return
        for f in files:
            inspect(f)
    else:
        inspect(target)


if __name__ == '__main__':
    main()
