# -*- coding: utf-8 -*-
"""규정 xlsx → 섹션 구조 추출 (0056 행단위 추출 결함 교정본, 260801).

사용:
  단건:  python scripts/reg_extract.py <xlsx> [--text]
  일괄:  python scripts/reg_extract.py --dir <폴더> --out <결과.json>

■ 왜 다시 만드는가
0056 은 시트를 '행 단위'로 읽어 텍스트화했다. 그런데 이 회사 규정 양식은
좌·우 2단 조판이라, 행 단위로 읽으면 왼쪽 단 문장과 오른쪽 단 문장이 한 줄에
섞인다(실측: A-6300 섹션 제목이 "1. 적용범위 4. 품질신고 및 포상절차"로 병합,
57종 중 53종 불량). 페이지·단(段)을 인식해 사람이 읽는 순서로 복원한다.

■ 템플릿 실측(정본 55종 전량 스캔, 260801)
  · 페이지 라벨 열 41 / 페이지 번호 열 47  — 55/55 일치
  · 우단 시작 열 30 — 53/55 (나머지 2종은 35·36 → 동적 탐지로 흡수)
  · 페이지 머리글 블록 = 페이지번호 행 기준 위로 4행(품질경영시스템·문서번호·재개정일자·개정번호·페이지)
  · 각 단의 '첫 열' = 본문 산문, 나머지 열 = 표(개정이력·승인란·흐름도)

■ 읽는 순서
  페이지1 좌단 → 페이지1 우단 → 페이지2 좌단 → … (원본 인쇄물과 동일)
  표(흐름도 등)는 산문에 섞지 않고 해당 섹션 끝에 [절차 흐름도]/[표] 블록으로 붙인다.

■ 값 처리
  · 숫자 서식이 통화/천단위면 그대로 포맷(포상금액 50000 → 50,000)
  · 엑셀 날짜 일련값은 (서두) 영역에서만 날짜로 환산 — 본문의 금액과 값 범위가
    겹치므로(50000 = 포상금액 vs 45821 = 2025-06-13) 범위 추정을 본문에 쓰면 위험하다.
    원본이 재,개정일자를 General 서식 숫자로 저장한 파일이 다수라 이 보정이 필요.
"""
import argparse
import datetime as _dt
import glob
import io
import json
import os
import re
import warnings

warnings.filterwarnings("ignore")
from openpyxl import load_workbook

PAGE_RE = re.compile(r"^\s*(\d+)\s*/\s*(\d+)\s*$")
# 최상위 표제: "1." "12." — "6.1" 같은 하위번호는 제외(뒤에 숫자가 오면 불일치)
HEAD_RE = re.compile(r"^\s*(\d{1,2})\s*\.(?!\d)\s*(.*)$")
# 2단계 표제: "8.1 운용 기획 및 관리" — "8.1.1"(3단계)은 제외.
# 품질환경매뉴얼처럼 한 장(章)이 3만자를 넘는 문서는 장 단위로만 쪼개면 KB 검색이
# 조각 하나를 통째로 반환해 어디를 볼지 못 짚어준다 → --sublevel 로 절(節)까지 분할.
SUB_HEAD_RE = re.compile(r"^\s*(\d{1,2})\.(\d{1,2})(?!\.?\d)\s*(.*)$")
HDR_LABELS = ("문서번호", "재,개정일자", "개정번호", "페이지", "품질경영시스템", "환경경영시스템")
PAGE_COL_HINT = 47
LABEL_COL_HINT = 41
JUNK = {"`", "´", "'", "-", "—", "ㆍ"}
# 2번째 이후 시트 처리: 부표/별표류만 규정 본문의 일부로 취급하고 양식 시트는 제외한다.
# (실측: F-2100 은 19시트 중 18시트가 양식 F2100-01~15 — 0056 은 이걸 본문에 섞어
#  본문자 7,172자로 부풀렸다. 양식은 forms/셀맵 계층 소관이라 규정 본문에 들어오면 안 된다.)
ANNEX_RE = re.compile(r"^\s*(부표|별표|별첨|붙임)")
# 점선 리더 줄("1. - - - - - - 적용범위") = 목차 또는 번호 표기 예시이지 실제 표제가 아니다.
# 표제로 오인하면 번호가 뒤로 튀어(6→3) 이후 정상 항들이 별첨으로 잘못 묶인다(실측 A-4101).
LEADER_RE = re.compile(r"(?:[-·.]\s+){3,}")
EPOCH = _dt.date(1899, 12, 30)  # 엑셀 1900 날짜계(윤년 버그 포함) 기준


def serial_to_date(n):
    try:
        return (EPOCH + _dt.timedelta(days=int(n))).isoformat()
    except Exception:
        return str(n)


def fmt_value(v, nf):
    """셀 값 → 표시 문자열(날짜/천단위 서식 반영)."""
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, bool):
        return "Y" if v else "N"
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v == int(v):
            v = int(v)
        if nf and ("#,##0" in nf or "₩" in nf):
            s = f"{v:,}"
            return f"₩{s}" if "₩" in nf else s
        return str(v)
    s = str(v).replace("\r\n", "\n").replace("\xa0", " ")
    s = re.sub(r"[ \t]+", " ", s)
    return s.strip()


def read_cells(ws):
    """병합 앵커에만 값이 있는 (행, 열, 표시문자열, 원시값) 목록."""
    non = set()
    for m in ws.merged_cells.ranges:
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                if (r, c) != (m.min_row, m.min_col):
                    non.add((r, c))
    out = []
    for r in range(1, (ws.max_row or 0) + 1):
        for c in range(1, (ws.max_column or 0) + 1):
            if (r, c) in non:
                continue
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            s = fmt_value(cell.value, cell.number_format)
            if s and s not in JUNK:
                out.append((r, c, s, cell.value))
    return out


def find_pages(cells):
    """페이지 경계 산출 → [(본문시작행, 본문끝행)].

    머리글 블록 높이를 상수로 두면 안 된다(실측: A-2100 (9)항이 한 행 차이로 잘림).
    페이지 번호 행 위쪽에서 '문서 번 호' 라벨 행을 찾아 그 행을 머리글 시작으로 삼는다.
    """
    marks = sorted(
        [r for r, c, s, _v in cells if PAGE_RE.match(s) and abs(c - PAGE_COL_HINT) <= 8]
    )
    if not marks:
        return []
    doc_no_rows = sorted(
        {r for r, c, s, _v in cells if c >= LABEL_COL_HINT and s.replace(" ", "") == "문서번호"}
    )

    def hdr_start(page_row):
        above = [h for h in doc_no_rows if h <= page_row]
        # 라벨을 못 찾으면 관측된 표준 높이(4행)로 후퇴
        return above[-1] if above else page_row - 3

    pages = []
    for i, r in enumerate(marks):
        body_start = r + 1
        body_end = (hdr_start(marks[i + 1]) - 1) if i + 1 < len(marks) else 10 ** 6
        pages.append((body_start, body_end))
    return pages


def detect_split(cells, pages):
    """우단 시작 열 — 본문 영역 사용 열의 최대 간격으로 탐지(기본 30)."""
    uniq = sorted({c for r, c, s, _v in cells if c < LABEL_COL_HINT
                   and any(a <= r <= b for a, b in pages)})
    best, at = 0, 30
    for i in range(len(uniq) - 1):
        gap = uniq[i + 1] - uniq[i]
        if 15 <= uniq[i + 1] <= 40 and gap > best:
            best, at = gap, uniq[i + 1]
    return at if best >= 4 else 30


def annex_sections(wb):
    """부표/별표 시트 → 표 형태 섹션. 페이지 머리글 템플릿이 없으므로 행 단위로 읽는다.

    '부표1_..._세로' / '_가로' 같은 레이아웃 변형은 한쪽이 다른 쪽에 완전히 포함될 때만
    버린다. 실측(M-1200)에서 두 변형이 서로 다른 문장을 각각 갖고 있어, '긴 쪽만 채택'
    으로 줄이면 7문장이 사라졌다 — 규정 본문은 중복을 감수하더라도 누락이 없어야 한다.
    """
    groups = {}
    for ws in wb.worksheets[1:]:
        if not ANNEX_RE.match(ws.title):
            continue
        by_row = {}
        for r, c, s, _v in read_cells(ws):
            by_row.setdefault(r, []).append((c, s))
        lines = [" | ".join(s for _c, s in sorted(by_row[r])) for r in sorted(by_row)]
        body = "\n".join(x for x in lines if x.strip())
        if not body:
            continue
        key = re.sub(r"[_\s]*(세로|가로|\(\d+\))\s*$", "", ws.title).strip()
        groups.setdefault(key, []).append((ws.title, body, wb.worksheets.index(ws)))

    kept = []
    for variants in groups.values():
        sets = [{l for l in b.split("\n") if l.strip()} for _t, b, _i in variants]
        for i, v in enumerate(variants):
            dup = any(
                i != j and (sets[i] < sets[j] or (sets[i] == sets[j] and j < i))
                for j in range(len(variants))
            )
            if not dup:
                kept.append(v)
    return [(f"[부표] {t}", b) for t, b, _i in sorted(kept, key=lambda x: x[2])]


def head_key(p, sublevel):
    """표제면 (대번호, 소번호, 제목) 반환, 아니면 None. sublevel 이면 'N.M'도 표제로."""
    if not p or LEADER_RE.search(p):
        return None
    if sublevel:
        m = SUB_HEAD_RE.match(p)
        if m:
            return int(m.group(1)), int(m.group(2)), m.group(3)
    m = HEAD_RE.match(p)
    return (int(m.group(1)), 0, m.group(2)) if m else None


def extract(path, sheet=None, sublevel=False):
    """sheet 지정 시 그 시트를 본문으로 삼는다(품질환경매뉴얼처럼 장(章)이 시트로 나뉜 문서)."""
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    cells = read_cells(ws)
    pages = find_pages(cells)
    # 페이지 머리글 템플릿이 없는 시트(예: 품질환경매뉴얼의 ZF CSR 매트릭스)는 경계가 안 잡혀
    # 본문 범위가 0이 된다 → 시트 전체를 1페이지로 보고 읽는다(실측: 378셀 35,352자가 통째로 누락됐음).
    if not pages:
        pages = [(1, 10 ** 6)]
    split = detect_split(cells, pages)

    # ── 메타 ── 머리글 라벨(열 41+) 오른쪽 값
    meta = {"sheet": ws.title, "pages": len(pages), "split_col": split,
            "doc_no": None, "rev_no": None, "rev_date": None, "title": None}
    # 문서 제목은 페이지 머리글에만 있다(본문 범위 밖) → meta 로 뽑아 (서두) 첫 줄에 싣는다.
    # 안 실으면 "신고포상" 같은 제목 검색이 KB 에서 0건이 된다(구본은 머리글을 본문에
    # 통째로 넣어 우연히 검색됐다). 머리글 행 중 라벨열 왼쪽의 시스템명 아닌 셀 = 제목.
    # 머리글 행 = '페이지 본문 범위의 여집합'. 고정 폭으로 잡으면 직전 페이지 본문 끝줄
    # (승인란 '일 자' 행)까지 삼켜 날짜가 제목 후보로 올라온다(실측 A-1100 → '2025-06-27').
    titles = [s for r, c, s, _v in cells
              if not any(a <= r <= b for a, b in pages) and c < LABEL_COL_HINT
              and s.replace(" ", "") not in ("품질경영시스템", "환경경영시스템")
              and not PAGE_RE.match(s) and len(s) >= 4]
    if titles:
        meta["title"] = max(set(titles), key=titles.count)
    for r, c, s, _v in cells:
        key = s.replace(" ", "")
        if c < LABEL_COL_HINT or key not in HDR_LABELS:
            continue
        vals = [(vv, ss) for rr, cc, ss, vv in cells if rr == r and cc > c]
        if not vals:
            continue
        raw, shown = vals[0]
        if key == "문서번호" and meta["doc_no"] is None:
            meta["doc_no"] = shown
        elif key == "개정번호" and meta["rev_no"] is None:
            meta["rev_no"] = shown
        elif key == "재,개정일자" and meta["rev_date"] is None:
            meta["rev_date"] = serial_to_date(raw) if isinstance(raw, int) else shown

    # ── 페이지 → 단 → 행 순서로 나열 (읽는 순서) ──
    # 항목 = (행, 산문열, [(열, 표시문자열, 원시값)…]) — 한 행의 셀을 쪼개지 않고 통째로 보관해야
    # (서두)의 "구 분 | 작 성 | 검 토" 같은 라벨+값 한 줄을 원형대로 복원할 수 있다.
    # 우단 상한은 두지 않는다 — 머리글 열(41·47)은 페이지 머리글 '행'에만 있고 그 행은
    # 이미 본문 범위 밖이다. 상한을 41로 막으면 승인/협의란 우측(43~52열)이 잘린다.
    BANDS = ((1, split - 1), (split, 10 ** 6))

    # 산문 열은 '문서 전체' 기준으로 밴드마다 하나씩 정한다. 페이지별로 정하면 흐름도가
    # 큰 페이지에서 흐름도 열이 산문 열로 뽑혀, 그 페이지의 표제가 표 셀로 흡수된다
    # (실측 L-2300: '5. 세부절차'가 4항 본문 속으로 사라짐).
    prose_cols = {}
    for lo, hi in BANDS:
        freq = {}
        for r, c, s, _v in cells:
            if lo <= c <= hi and any(a <= r <= b for a, b in pages):
                freq[c] = freq.get(c, 0) + 1
        if freq:
            prose_cols[(lo, hi)] = min(sorted(freq, key=lambda c: (-freq[c], c))[:2])

    stream = []
    for body_start, body_end in pages:
        page_cells = [(r, c, s, v) for r, c, s, v in cells if body_start <= r <= body_end]
        for lo, hi in BANDS:
            band = [(r, c, s, v) for r, c, s, v in page_cells if lo <= c <= hi]
            if not band or (lo, hi) not in prose_cols:
                continue
            prose_col = prose_cols[(lo, hi)]
            by_row = {}
            for r, c, s, v in band:
                by_row.setdefault(r, []).append((c, s, v))
            for r in sorted(by_row):
                stream.append((r, prose_col, sorted(by_row[r])))

    def split_row(item):
        """한 행을 (산문, 표셀)로 가른다.

        기준 열만 산문으로 보면, 같은 규정 안에서 문단 상자가 다른 열에서 시작하는 페이지의
        표제가 통째로 사라진다(실측 A-4200: '6. 관련양식'·'7. 부표'). 반대로 행 전체를
        산문으로 보면 흐름도·승인란이 문장에 섞인다. 그래서 둘을 합친다 —
        밴드 안에서 그 행에 셀이 하나뿐이면 산문(표는 한 행에 여러 칸을 쓴다).
        오분류 비용이 비대칭이라 산문 쪽으로 기운다: 표가 문장이 되면 읽기 불편할 뿐이지만,
        표제가 표로 빨려가면 그 절(節)이 통째로 사라진다.
        """
        _r, pc, items = item
        if len(items) == 1:
            return items[0][1], []
        return (" ".join(s for c, s, _v in items if c == pc),
                [(s, v) for c, s, v in items if c != pc])

    def prose_of(item):
        return split_row(item)[0]

    # ── 첫 표제 위치 = (서두)/본문 경계 ──
    first_head = len(stream)
    for i, item in enumerate(stream):
        if head_key(prose_of(item), sublevel):
            first_head = i
            break

    # (서두) = 경계 앞 전부 — 행 단위로 셀을 열 순서대로 이어 표 형태 복원
    pre_lines = []
    for _r, _pc, items in stream[:first_head]:
        vals = [serial_to_date(v) if isinstance(v, int) and 20000 <= v <= 60000 else s
                for _c, s, v in items]
        pre_lines.append(" | ".join(vals))
    head_line = " · ".join(x for x in [
        meta["title"],
        f"문서번호 {meta['doc_no']}" if meta["doc_no"] else None,
        f"개정 REV.{meta['rev_no']}" if meta["rev_no"] is not None else None,
        f"재·개정일 {meta['rev_date']}" if meta["rev_date"] else None,
    ] if x)
    preamble = "\n".join([head_line] + [x for x in pre_lines if x.strip()]) if head_line \
        else "\n".join(x for x in pre_lines if x.strip())

    # ── 본문 섹션 분해 ──
    # 최상위 번호가 뒤로 돌아가면(13→1) 그 지점부터 부표/별첨의 자체 번호 체계다(실측 14종).
    # 제목만 보면 "1. 적용범위"가 두 번 나와 중복처럼 보이므로, 직전에 나온 부표 표제를
    # 문맥 접두어로 붙여 구분한다. 원문 번호는 그대로 두고 접두어만 추가(원문 훼손 없음).
    sections = []
    cur_title, cur_body, cur_tbl = None, [], []
    last_key, ctx, ctx_n = (0, 0), "", 0
    annex_label = None

    def flush():
        if cur_title is None:
            return
        body = "\n".join(x for x in cur_body if x)
        if cur_tbl:
            flow = any(s.replace(" ", "").startswith("책임(") for row in cur_tbl for s, _v in row)
            lines = [" | ".join(s for s, _v in row) for row in cur_tbl]
            label = "[절차 흐름도]" if flow else "[표]"
            body = f"{body}\n\n{label}\n" + "\n".join(lines)
        sections.append((cur_title, body.strip()))

    for item in stream[first_head:]:
        p, tbl = split_row(item)
        if p and ANNEX_RE.match(p):
            annex_label = re.sub(r"\s+", " ", p)[:24]
        hk = head_key(p, sublevel)
        if hk and len(p) <= 40:
            key = (hk[0], hk[1])
            if key < last_key:  # 번호 재시작 = 부표/별첨 진입
                ctx_n += 1
                ctx = f"[{annex_label}] " if annex_label else f"[별첨{ctx_n}] "
                annex_label = None
            last_key = key
            flush()
            cur_title, cur_body, cur_tbl = ctx + p, [], []
        elif hk:
            # 표제 + 본문이 한 셀에 붙은 경우 — 앞부분을 제목으로 절단
            flush()
            num = f"{hk[0]}.{hk[1]}" if hk[1] else f"{hk[0]}."
            cut = re.split(r"(?<=[가-힣])\s(?=[(0-9가-힣])", hk[2], 1)
            cur_title = f"{num} {cut[0]}"[:40]
            cur_body, cur_tbl = [p], []
        elif p:
            cur_body.append(p)
        if tbl and cur_title:
            cur_tbl.append(tbl)
    flush()

    # 본문이 빈 섹션도 남긴다. 목차 점선 리더는 LEADER_RE 가 이미 표제에서 배제하므로,
    # 여기 남는 빈 섹션은 '장 표제'처럼 제목만 있는 실제 표제다(실측: --sublevel 사용 시
    # "8. 운용"이 8.1 직전에 빈 섹션이 되는데, 버리면 장 제목 텍스트가 통째로 사라진다).
    meta["empty_sections"] = [t for t, b in sections if not b.strip()]

    sections.extend(annex_sections(wb))
    meta["annex_sheets"] = [t for t, _b in sections if t.startswith("[부표]")]
    meta["skipped_sheets"] = [w.title for w in wb.worksheets[1:] if not ANNEX_RE.match(w.title)]
    return {"file": os.path.basename(path), "meta": meta, "preamble": preamble,
            "sections": [{"title": t, "body": b} for t, b in sections]}


def to_text(doc):
    buf = io.StringIO()
    m = doc["meta"]
    buf.write(f"■ {doc['file']}\n")
    buf.write(f"  문서번호 {m['doc_no']} · 개정 {m['rev_no']} · 일자 {m['rev_date']} "
              f"· {m['pages']}페이지 · 우단열 {m['split_col']}\n\n")
    if doc["preamble"]:
        buf.write("───── (서두) ─────\n" + doc["preamble"] + "\n\n")
    for s in doc["sections"]:
        buf.write(f"───── {s['title']} ─────\n{s['body']}\n\n")
    return buf.getvalue()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?")
    ap.add_argument("--dir")
    ap.add_argument("--out")
    ap.add_argument("--text", action="store_true")
    ap.add_argument("--sheet")
    ap.add_argument("--sheets", action="store_true", help="시트 목록만 출력")
    ap.add_argument("--sublevel", action="store_true", help="'8.1' 절 단위까지 분할")
    a = ap.parse_args()

    if a.sheets:
        for i, n in enumerate(load_workbook(a.xlsx, data_only=True).sheetnames):
            print(f"  [{i}] {n}")
        return

    if a.dir:
        docs = []
        for p in sorted(glob.glob(os.path.join(a.dir, "*.xlsx"))):
            if os.path.basename(p).startswith("~$"):
                continue
            try:
                docs.append(extract(p))
            except Exception as e:
                docs.append({"file": os.path.basename(p), "error": repr(e)})
        if a.out:
            with io.open(a.out, "w", encoding="utf-8") as fp:
                json.dump(docs, fp, ensure_ascii=False, indent=1)
        ok = [d for d in docs if "error" not in d]
        print(f"추출 {len(ok)}/{len(docs)}종 · 섹션 합계 {sum(len(d['sections']) for d in ok)}")
        for d in docs:
            if "error" in d:
                print(f"  ⚠ {d['file']}: {d['error']}")
        return

    doc = extract(a.xlsx, a.sheet, a.sublevel)
    print(to_text(doc) if a.text else json.dumps(doc, ensure_ascii=False, indent=1))


main()
